"""Unit tests for modules/grading/result_builder.py.

PR-10 hardening coverage for exporter and mode-summary helpers.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import openpyxl

from modules.grading.result_builder import (
    AttemptResultData,
    ExamResultExporter,
    ModeSummaryBuilder,
    QuestionResultRow,
)


def _sample_data(mode: str = "PRACTICE") -> AttemptResultData:
    now = datetime.now(timezone.utc)
    return AttemptResultData(
        submitter_name="Nguyen Van A",
        submitter_id="S001",
        quiz_title="Quiz Demo",
        mode=mode,
        started_at=now,
        submitted_at=now,
        duration_seconds=125,
        score=1.5,
        max_score=2.0,
        correct_count=1,
        incorrect_count=0,
        skipped_count=1,
        questions=[
            QuestionResultRow(
                order=1,
                question_text="2 + 2 = ?",
                answer_text="B",
                is_correct=True,
                score_awarded=1.0,
                max_score=1.0,
                question_code="Q1",
                correct_answer_display="B",
            ),
            QuestionResultRow(
                order=2,
                question_text="Thủ đô Pháp?",
                answer_text="Bỏ qua",
                is_correct=None,
                score_awarded=0.5,
                max_score=1.0,
                question_code="Q2",
                correct_answer_display="Paris",
            ),
        ],
    )


def test_build_excel_returns_valid_workbook_bytes() -> None:
    exporter = ExamResultExporter()
    excel_bytes = exporter.build_excel(_sample_data())

    assert isinstance(excel_bytes, (bytes, bytearray))
    assert len(excel_bytes) > 0

    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb["Kết quả làm bài"]

    assert ws["A1"].value == "Họ và tên:"
    assert ws["B1"].value == "Nguyen Van A"
    assert ws["A13"].value == "STT"
    assert ws["B13"].value == "Nội dung câu hỏi"


def test_mode_summary_exam_is_minimal() -> None:
    html = ModeSummaryBuilder.build_html(_sample_data(mode="EXAM"))

    assert "Đã hoàn thành" in html
    assert "Kết quả đã được ghi nhận" in html
    assert "Điểm:" not in html


def test_mode_summary_practice_contains_totals() -> None:
    html = ModeSummaryBuilder.build_html(_sample_data(mode="PRACTICE"))

    assert "Kết quả" in html
    assert "Điểm:" in html
    assert "Đúng:" in html
    assert "Bỏ qua:" in html


def test_mode_summary_study_uses_summary_layout() -> None:
    html = ModeSummaryBuilder.build_html(_sample_data(mode="STUDY"))

    assert "Kết quả" in html
    assert "Học tập" in html


def test_duration_format_variants() -> None:
    assert ModeSummaryBuilder._format_duration(45) == "45 giây"
    assert ModeSummaryBuilder._format_duration(125) == "2 phút 5 giây"
    assert ModeSummaryBuilder._format_duration(3725) == "1 giờ 2 phút 5 giây"


def test_exporter_result_text_variants() -> None:
    assert ExamResultExporter._result_text(True) == "Đúng ✓"
    assert ExamResultExporter._result_text(False) == "Sai ✗"
    assert ExamResultExporter._result_text(None) == "—"


def test_mode_label_passthrough_for_unknown_mode() -> None:
    assert ExamResultExporter._mode_label("CUSTOM") == "CUSTOM"
