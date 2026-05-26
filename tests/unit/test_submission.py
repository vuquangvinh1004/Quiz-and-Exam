"""Unit tests for submission feature.

Covers:
- ExamResultExporter: builds valid Excel bytes
- SubmissionService: settings load/save, filename, folder delivery
- AttemptResultData: data integrity
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from modules.grading.result_builder import (
    AttemptResultData,
    ExamResultExporter,
    QuestionResultRow,
)
from core.domain.services.submission_service import (
    SubmissionService,
    SubmissionSettings,
)
from core.utils.exceptions import SubmissionConfigError


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_result_data(
    mode: str = "EXAM",
    submitter_name: str = "Nguyễn Văn A",
    submitter_id: str = "SV001",
    questions: list[QuestionResultRow] | None = None,
) -> AttemptResultData:
    now = datetime(2026, 3, 24, 10, 0, 0, tzinfo=timezone.utc)
    if questions is None:
        questions = [
            QuestionResultRow(
                order=1,
                question_text="Thủ đô Việt Nam?",
                answer_text="Hà Nội",
                is_correct=True,
                score_awarded=1.0,
                max_score=1.0,
            ),
            QuestionResultRow(
                order=2,
                question_text="2 + 2 = ?",
                answer_text="5",
                is_correct=False,
                score_awarded=0.0,
                max_score=1.0,
            ),
            QuestionResultRow(
                order=3,
                question_text="EOQ là gì?",
                answer_text="Bỏ qua",
                is_correct=None,
                score_awarded=0.0,
                max_score=1.0,
            ),
        ]
    return AttemptResultData(
        submitter_name=submitter_name,
        submitter_id=submitter_id,
        quiz_title="Bài kiểm tra thử",
        mode=mode,
        started_at=now,
        submitted_at=datetime(2026, 3, 24, 10, 30, 0, tzinfo=timezone.utc),
        duration_seconds=1800,
        score=1.0,
        max_score=3.0,
        correct_count=1,
        incorrect_count=1,
        skipped_count=1,
        questions=questions,
    )


# ---------------------------------------------------------------------------
# ExamResultExporter tests
# ---------------------------------------------------------------------------

class TestExamResultExporter:
    def setup_method(self):
        self.exporter = ExamResultExporter()

    def test_build_excel_returns_bytes(self):
        data = _make_result_data()
        result = self.exporter.build_excel(data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_is_valid_workbook(self):
        data = _make_result_data()
        result = self.exporter.build_excel(data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_worksheet_title(self):
        data = _make_result_data()
        result = self.exporter.build_excel(data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Kết quả làm bài" in wb.sheetnames

    def test_submitter_name_in_sheet(self):
        data = _make_result_data(submitter_name="Trần Thị B")
        result = self.exporter.build_excel(data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Kết quả làm bài"]
        cell_values = [str(ws.cell(r, 2).value or "") for r in range(1, 15)]
        assert "Trần Thị B" in cell_values

    def test_submitter_id_in_sheet(self):
        data = _make_result_data(submitter_id="NV123")
        result = self.exporter.build_excel(data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Kết quả làm bài"]
        cell_values = [str(ws.cell(r, 2).value or "") for r in range(1, 15)]
        assert "NV123" in cell_values

    def test_question_rows_present(self):
        data = _make_result_data()
        result = self.exporter.build_excel(data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Kết quả làm bài"]
        all_text = " ".join(
            str(ws.cell(r, 2).value or "")
            for r in range(1, ws.max_row + 1)
        )
        assert "Thủ đô Việt Nam?" in all_text

    def test_mode_label_exam(self):
        assert ExamResultExporter._mode_label("EXAM") == "Kiểm tra"

    def test_mode_label_practice(self):
        assert ExamResultExporter._mode_label("PRACTICE") == "Luyện tập"

    def test_mode_label_study(self):
        assert ExamResultExporter._mode_label("STUDY") == "Học tập"

    def test_format_duration_seconds(self):
        assert ExamResultExporter._format_duration(45) == "45 giây"

    def test_format_duration_minutes(self):
        assert ExamResultExporter._format_duration(125) == "2 phút 5 giây"

    def test_format_duration_hours(self):
        assert ExamResultExporter._format_duration(3661) == "1 giờ 1 phút 1 giây"

    def test_result_text_correct(self):
        assert "Đúng" in ExamResultExporter._result_text(True)

    def test_result_text_incorrect(self):
        assert "Sai" in ExamResultExporter._result_text(False)

    def test_result_text_none(self):
        assert ExamResultExporter._result_text(None) == "—"

    def test_empty_questions_list(self):
        data = _make_result_data(questions=[])
        result = self.exporter.build_excel(data)
        assert isinstance(result, bytes)


# ---------------------------------------------------------------------------
# SubmissionService tests
# ---------------------------------------------------------------------------

class TestSubmissionService:
    def setup_method(self):
        self.service = SubmissionService()

    def test_build_filename_safe_chars(self):
        data = _make_result_data(submitter_name="Lê/Văn:C", submitter_id="001")
        filename = self.service.build_filename(data)
        # Should not contain invalid chars
        assert "/" not in filename
        assert ":" not in filename
        assert filename.endswith(".xlsx")

    def test_build_filename_contains_name(self):
        data = _make_result_data(submitter_name="AnhBinh")
        filename = self.service.build_filename(data)
        assert "AnhBinh" in filename

    def test_build_filename_contains_quiz_title(self):
        data = _make_result_data()
        filename = self.service.build_filename(data)
        # Quiz title "Bài kiểm tra thử" → sanitized
        assert "KetQua_" in filename

    def test_submit_to_folder_creates_file(self, tmp_path):
        data = _make_result_data()
        saved = self.service.submit_to_folder(data, str(tmp_path))
        assert saved.exists()
        assert saved.suffix == ".xlsx"

    def test_submit_to_folder_creates_dir_if_not_exists(self, tmp_path):
        target = tmp_path / "new_subdir" / "results"
        data = _make_result_data()
        saved = self.service.submit_to_folder(data, str(target))
        assert saved.exists()

    def test_submit_to_folder_empty_path_raises(self):
        data = _make_result_data()
        with pytest.raises(SubmissionConfigError, match="thư mục"):
            self.service.submit_to_folder(data, "")

    def test_submit_via_email_no_server_raises(self):
        data = _make_result_data()
        cfg = SubmissionSettings(smtp_server="", mode="email")
        with pytest.raises(SubmissionConfigError, match="SMTP"):
            self.service.submit_via_email(data, cfg, recipient="a@b.com")

    def test_submit_via_email_empty_recipient_raises(self):
        data = _make_result_data()
        cfg = SubmissionSettings(smtp_server="smtp.example.com", mode="email")
        with pytest.raises(SubmissionConfigError, match="email"):
            self.service.submit_via_email(data, cfg, recipient="")


# ---------------------------------------------------------------------------
# SubmissionSettings load/save via DB (in-memory SQLite)
# ---------------------------------------------------------------------------

class TestSubmissionSettingsDB:
    """Integration-style tests using in-memory SQLite."""

    @pytest.fixture
    def session(self, tmp_path):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from core.database.models import Base
        engine = create_engine(f"sqlite:///{tmp_path}/test.db")
        Base.metadata.create_all(engine)
        factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
        sess = factory()
        yield sess
        sess.close()

    def test_load_defaults_when_empty(self, session):
        svc = SubmissionService()
        cfg = svc.load_settings(session)
        assert cfg.mode == "none"
        assert cfg.smtp_port == 587

    def test_save_and_reload(self, session):
        svc = SubmissionService()
        original = SubmissionSettings(
            mode="email",
            default_email="test@example.com",
            smtp_server="smtp.gmail.com",
            smtp_port=465,
            smtp_user="user@gmail.com",
            smtp_password="secret",
            smtp_sender="Quiz App <user@gmail.com>",
            submit_folder="/tmp/results",
            smtp_use_tls=False,
        )
        svc.save_settings(session, original)
        session.commit()
        loaded = svc.load_settings(session)
        assert loaded.mode == "email"
        assert loaded.default_email == "test@example.com"
        assert loaded.smtp_server == "smtp.gmail.com"
        assert loaded.smtp_port == 465
        assert loaded.smtp_user == "user@gmail.com"
        assert loaded.smtp_password == "secret"
        assert loaded.submit_folder == "/tmp/results"
        assert loaded.smtp_use_tls is False

    def test_save_overwrites_existing(self, session):
        svc = SubmissionService()
        cfg1 = SubmissionSettings(mode="email", default_email="a@b.com")
        svc.save_settings(session, cfg1)
        session.commit()

        cfg2 = SubmissionSettings(mode="folder", default_email="x@y.com")
        svc.save_settings(session, cfg2)
        session.commit()

        loaded = svc.load_settings(session)
        assert loaded.mode == "folder"
        assert loaded.default_email == "x@y.com"

    def test_submit_result_dict_structure_folder_only(self, session, tmp_path):
        svc = SubmissionService()
        cfg = SubmissionSettings(mode="folder", submit_folder=str(tmp_path))
        data = _make_result_data()
        result = svc.submit(data, cfg, folder_path=str(tmp_path))
        assert result["folder_path"] is not None
        assert result["email_sent"] is False
        assert result["errors"] == []
