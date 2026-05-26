"""Build Excel result reports for quiz attempt submission.

This module defines AttemptResultData (a plain dataclass) and
ExamResultExporter which converts it to an openpyxl workbook (bytes).
ModeSummaryBuilder builds mode-appropriate HTML summary strings for
the end-of-quiz dialogs.

Business rules (ARCHITECTURE §7):
- Exam mode  : show all answersquestions with correct/incorrect flags (for examiner).
- Practice   : show correct/incorrect summary; no per-question correctness override.
- Study mode : per-question feedback is already in the data.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# Data transfer objects
# ---------------------------------------------------------------------------

@dataclass
class QuestionResultRow:
    """Per-question result row included in the submission report."""

    order: int
    question_text: str
    answer_text: str            # human-readable answer (or "Bỏ qua" if unanswered)
    is_correct: Optional[bool]  # None = not graded / pending
    score_awarded: float
    max_score: float
    question_code: Optional[str] = None      # original question_code from bank
    correct_answer_display: str = ""         # human-readable correct answer(s)


@dataclass
class AttemptResultData:
    """All data needed to render a submission report."""

    submitter_name: str
    submitter_id: str
    quiz_title: str
    mode: str                   # "EXAM" | "PRACTICE" | "STUDY"
    started_at: datetime
    submitted_at: datetime
    duration_seconds: int
    score: float
    max_score: float
    correct_count: int
    incorrect_count: int
    skipped_count: int
    questions: list[QuestionResultRow] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Excel exporter
# ---------------------------------------------------------------------------

class ExamResultExporter:
    """Converts AttemptResultData into an Excel workbook returned as bytes."""

    # Colour palette
    _BLUE_FILL = PatternFill("solid", fgColor="4472C4")
    _ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
    _CORRECT_FILL = PatternFill("solid", fgColor="C6EFCE")
    _INCORRECT_FILL = PatternFill("solid", fgColor="FFC7CE")
    _INFO_LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")

    def build_excel(self, data: AttemptResultData) -> bytes:
        """Return the workbook as raw bytes suitable for writing to a file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kết quả làm bài"

        row = self._write_info_block(ws, data, start_row=1)
        row += 1  # blank separator
        row = self._write_table(ws, data, start_row=row)

        self._set_column_widths(ws)
        ws.freeze_panes = f"A{row - len(data.questions)}"  # freeze above data

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _write_info_block(self, ws, data: AttemptResultData, start_row: int) -> int:
        """Write the information header block. Returns next available row."""
        info_rows = [
            ("Họ và tên:",       data.submitter_name),
            ("ID / Mã số:",       data.submitter_id),
            ("Bài kiểm tra:",    data.quiz_title),
            ("Chế độ:",          self._mode_label(data.mode)),
            ("Thời gian bắt đầu:", data.started_at.strftime("%d/%m/%Y %H:%M:%S")),
            ("Thời gian nộp bài:", data.submitted_at.strftime("%d/%m/%Y %H:%M:%S")),
            ("Thời gian làm bài:", self._format_duration(data.duration_seconds)),
            ("Điểm số:",          f"{data.score:.2f} / {data.max_score:.2f}"),
            ("Số câu đúng:",      str(data.correct_count)),
            ("Số câu sai:",       str(data.incorrect_count)),
            ("Số câu bỏ qua:",    str(data.skipped_count)),
        ]
        row = start_row
        for label, value in info_rows:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.fill = self._INFO_LABEL_FILL
            ws.cell(row=row, column=2, value=value)
            row += 1
        return row

    def _write_table(self, ws, data: AttemptResultData, start_row: int) -> int:
        """Write the per-question table. Returns next available row."""
        headers = ["STT", "Nội dung câu hỏi", "Câu trả lời", "Kết quả", "Điểm", "Điểm tối đa"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self._BLUE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row = start_row + 1

        for i, q in enumerate(data.questions):
            bg_fill = self._ALT_FILL if i % 2 == 1 else None
            result_text = self._result_text(q.is_correct)
            result_fill = (
                self._CORRECT_FILL if q.is_correct is True
                else self._INCORRECT_FILL if q.is_correct is False
                else None
            )
            row_data = [
                q.order,
                q.question_text,
                q.answer_text,
                result_text,
                f"{q.score_awarded:.2f}",
                f"{q.max_score:.2f}",
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if bg_fill:
                    cell.fill = bg_fill
                if col == 4 and result_fill:
                    cell.fill = result_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        return row

    @staticmethod
    def _set_column_widths(ws) -> None:
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 52
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12

    @staticmethod
    def _mode_label(mode: str) -> str:
        return {"EXAM": "Kiểm tra", "PRACTICE": "Luyện tập", "STUDY": "Học tập"}.get(mode, mode)

    @staticmethod
    def _format_duration(seconds: int) -> str:
        if seconds < 60:
            return f"{seconds} giây"
        minutes, secs = divmod(seconds, 60)
        if minutes < 60:
            return f"{minutes} phút {secs} giây"
        hours, mins = divmod(minutes, 60)
        return f"{hours} giờ {mins} phút {secs} giây"

    @staticmethod
    def _result_text(is_correct: Optional[bool]) -> str:
        if is_correct is True:
            return "Đúng ✓"
        if is_correct is False:
            return "Sai ✗"
        return "—"


# ---------------------------------------------------------------------------
# Mode-aware summary builder
# ---------------------------------------------------------------------------

class ModeSummaryBuilder:
    """Builds mode-appropriate HTML summary for end-of-quiz display.

    Rules (ARCHITECTURE §7):
    - EXAM     : minimal – completion notice only.
    - PRACTICE : summary – correct / wrong / skipped / score totals.
    - STUDY    : per_question – per-question detail was shown inline;
                 display same summary as PRACTICE.
    """

    _MODE_LABELS: dict[str, str] = {
        "EXAM": "Kiểm tra",
        "PRACTICE": "Luyện tập",
        "STUDY": "Học tập",
    }

    @classmethod
    def build_html(cls, data: "AttemptResultData") -> str:
        """Return an HTML string suitable for a QLabel or QMessageBox.

        Parameters
        ----------
        data:
            Fully populated AttemptResultData.

        Returns
        -------
        str
            Rich-text HTML fragment.
        """
        mode = data.mode
        label = cls._MODE_LABELS.get(mode, mode)
        pct = (data.score / data.max_score * 100) if data.max_score else 0.0

        if mode == "EXAM":
            return (
                f"<b>Đã hoàn thành – {label}</b><br><br>"
                f"Bài: {data.quiz_title}<br>"
                f"Kết quả đã được ghi nhận."
            )

        # PRACTICE and STUDY: show full summary
        duration_str = cls._format_duration(data.duration_seconds)
        return (
            f"<b>Kết quả – {label}</b><br><br>"
            f"<b>Bài:</b> {data.quiz_title}<br>"
            f"<b>Điểm:</b> {data.score:.2f} / {data.max_score:.2f}"
            f" ({pct:.1f}%)<br>"
            f"<b>Đúng:</b> {data.correct_count}"
            f"  &nbsp; <b>Sai:</b> {data.incorrect_count}"
            f"  &nbsp; <b>Bỏ qua:</b> {data.skipped_count}<br>"
            f"<b>Thời gian làm bài:</b> {duration_str}"
        )

    @staticmethod
    def _format_duration(seconds: int) -> str:
        if seconds < 60:
            return f"{seconds} giây"
        minutes, secs = divmod(seconds, 60)
        if minutes < 60:
            return f"{minutes} phút {secs} giây"
        hours, mins = divmod(minutes, 60)
        return f"{hours} giờ {mins} phút {secs} giây"
