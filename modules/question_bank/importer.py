"""CSV and Excel question file parser with per-row validation.

Follows QUIZ_APP_IMPORT_FORMAT.md strictly:
- Reads by column name, not position.
- Strict mode: any ERROR-level issue sets has_errors=True and blocks commit.
- Delimiter for multi-value fields: ||
  - Canonical placeholder for blank questions: [[blank]]
  - Legacy placeholder ________ is still accepted with a deprecation warning.
  - Question types supported: MC, MA, BLANK, TF, SA, ES.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from core.utils.constants import (
    BLANK_PLACEHOLDER,
    LEGACY_BLANK_PLACEHOLDER,
    DEFAULT_DIFFICULTY,
    DEFAULT_SCORE,
    DEFAULT_STATUS,
    MULTI_VALUE_DELIMITER,
    QUESTION_TYPE_IMPORT_MAP,
    VALID_OPTION_LABELS,
    Difficulty,
    QuestionStatus,
    QuestionType,
)
from core.utils.validators import (
    count_blank_placeholders,
    validate_correct_answers_for_type,
)
from core.utils.logger import get_logger

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = frozenset({"question_text", "question_type", "correct_answers"})

# Maps option label → import column name
OPTION_COLUMNS: dict[str, str] = {
    "A": "option_a",
    "B": "option_b",
    "C": "option_c",
    "D": "option_d",
    "E": "option_e",
    "F": "option_f",
}

VALID_BOOL_TRUE = frozenset({"true", "1", "yes", "y"})
VALID_BOOL_FALSE = frozenset({"false", "0", "no", "n"})
DEFAULT_SOFT_ROW_LIMIT = 12_000
DEFAULT_HARD_ROW_LIMIT = 30_000
DEFAULT_SOFT_FILE_SIZE_BYTES = 3 * 1024 * 1024
DEFAULT_HARD_FILE_SIZE_BYTES = 8 * 1024 * 1024

logger = get_logger(__name__)


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class ImportIssue:
    """A single validation finding for a row."""
    row: int          # 1-based (0 = file-level or header)
    severity: str     # 'ERROR', 'WARNING', 'INFO'
    message: str
    column: str = ""


@dataclass
class ParsedQuestion:
    """Normalized, validated data for one successfully parsed row."""
    row_number: int
    question_code: str | None
    question_text: str
    question_type: QuestionType
    category: str | None
    difficulty: str
    score: float
    hint: str | None
    explanation: str | None
    options: dict[str, str]          # {A: text, B: text, ...} – empty for BLANK/SA/ES
    correct_answers: list[str]       # ['A'] for MC/TF; ['A','C'] for MA; ['Hà Nội'] for BLANK/SA/ES
    status: str
    tags: list[str]
    case_sensitive: bool
    trim_whitespace: bool


@dataclass
class ParseResult:
    """Result of parsing an entire import file."""
    total_rows: int = 0
    parsed_questions: list[ParsedQuestion] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return any(i.severity == "ERROR" for i in self.issues)

    @property
    def error_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "ERROR")

    @property
    def warning_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "WARNING")

    @property
    def info_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == "INFO")


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

class QuestionFileParser:
    """Parse .csv or .xlsx question import files into ParseResult."""

    def __init__(
        self,
        *,
        soft_row_limit: int = DEFAULT_SOFT_ROW_LIMIT,
        hard_row_limit: int = DEFAULT_HARD_ROW_LIMIT,
        soft_file_size_bytes: int = DEFAULT_SOFT_FILE_SIZE_BYTES,
        hard_file_size_bytes: int = DEFAULT_HARD_FILE_SIZE_BYTES,
    ) -> None:
        self._soft_row_limit = soft_row_limit
        self._hard_row_limit = hard_row_limit
        self._soft_file_size_bytes = soft_file_size_bytes
        self._hard_file_size_bytes = hard_file_size_bytes

    def parse_file(self, path: Path) -> ParseResult:
        """Entry point: dispatch to CSV or XLSX reader based on file suffix."""
        file_size_issues, hard_stop_result = self._check_file_size_budget(path)
        if hard_stop_result is not None:
            return hard_stop_result

        suffix = path.suffix.lower()
        if suffix == ".csv":
            result = self._parse_csv(path)
        elif suffix in (".xlsx", ".xlsm"):
            result = self._parse_xlsx(path)
        else:
            result = ParseResult()
            result.issues.append(ImportIssue(
                row=0, severity="ERROR",
                message=(
                    f"Định dạng file không được hỗ trợ: '{suffix}'. "
                    "Chỉ chấp nhận .csv và .xlsx."
                ),
            ))
        result.issues.extend(file_size_issues)
        result.issues.sort(key=lambda i: (i.row, i.severity))
        return result

    # -----------------------------------------------------------------------
    # File readers
    # -----------------------------------------------------------------------

    def _parse_csv(self, path: Path) -> ParseResult:
        try:
            raw = path.read_bytes()
            text = raw.decode("utf-8-sig")   # handles BOM transparently
            reader = csv.DictReader(io.StringIO(text))
            headers: list[str] = list(reader.fieldnames or [])
        except Exception as exc:
            result = ParseResult()
            result.issues.append(ImportIssue(
                row=0, severity="ERROR",
                message=f"Không thể đọc file CSV: {exc}",
            ))
            return result
        return self._process_rows(headers, reader)

    def _parse_xlsx(self, path: Path) -> ParseResult:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb["questions"] if "questions" in wb.sheetnames else wb.active
            row_iter = ws.iter_rows()
        except Exception as exc:
            result = ParseResult()
            result.issues.append(ImportIssue(
                row=0, severity="ERROR",
                message=f"Không thể đọc file Excel: {exc}",
            ))
            return result

        try:
            header_row = next(row_iter)
        except StopIteration:
            result = ParseResult()
            result.issues.append(ImportIssue(
                row=0, severity="ERROR", message="File Excel trống.",
            ))
            return result

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in header_row
        ]

        def _iter_dict_rows() -> Iterable[dict[str, Any]]:
            for row in row_iter:
                row_dict: dict[str, Any] = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        val = cell.value
                        row_dict[headers[col_idx]] = (
                            str(val).strip() if val is not None else ""
                        )
                if any(v for v in row_dict.values()):
                    yield row_dict

        return self._process_rows(headers, _iter_dict_rows())

    # -----------------------------------------------------------------------
    # Core processing
    # -----------------------------------------------------------------------

    def _process_rows(
        self, headers: list[str], rows: Iterable[dict[str, Any]]
    ) -> ParseResult:
        result = ParseResult()

        # Header validation first
        header_issues = self._validate_headers(headers)
        result.issues.extend(header_issues)
        if any(i.severity == "ERROR" for i in header_issues):
            return result  # Cannot continue without required columns

        soft_limit_warning_added = False
        for idx, row in enumerate(rows, start=1):
            if idx > self._hard_row_limit:
                result.issues.append(ImportIssue(
                    row=0,
                    severity="ERROR",
                    column="file",
                    message=(
                        f"File có hơn {self._hard_row_limit} dòng dữ liệu. "
                        "Vượt ngưỡng an toàn import hiện tại; vui lòng chia nhỏ file."
                    ),
                ))
                logger.warning(
                    "event=import_row_limit_exceeded rows=%s hard_limit=%s"
                    % (idx, self._hard_row_limit)
                )
                break
            if idx > self._soft_row_limit and not soft_limit_warning_added:
                result.issues.append(ImportIssue(
                    row=0,
                    severity="WARNING",
                    column="file",
                    message=(
                        f"File lớn ({idx}+ dòng). Preview/import có thể chậm hơn bình thường; "
                        "hãy ưu tiên chạy trên máy ổn định và tránh đóng ứng dụng giữa chừng."
                    ),
                ))
                soft_limit_warning_added = True
            result.total_rows += 1
            # Skip completely blank rows
            if not any(str(v).strip() for v in row.values()):
                result.total_rows -= 1
                continue
            parsed, row_issues = self._parse_row(idx, row)
            result.issues.extend(row_issues)
            if parsed is not None:
                result.parsed_questions.append(parsed)

        return result

    def _check_file_size_budget(
        self, path: Path
    ) -> tuple[list[ImportIssue], ParseResult | None]:
        try:
            file_size = path.stat().st_size
        except OSError:
            return [], None

        issues: list[ImportIssue] = []

        if file_size > self._hard_file_size_bytes:
            result = ParseResult()
            result.issues.append(ImportIssue(
                row=0,
                severity="ERROR",
                column="file",
                message=(
                    f"File quá lớn ({_format_size(file_size)}). "
                    "Vượt ngưỡng an toàn import hiện tại; vui lòng chia nhỏ file."
                ),
            ))
            logger.warning(
                "event=import_file_too_large path=%s size=%s hard_limit=%s"
                % (path.name, file_size, self._hard_file_size_bytes)
            )
            return [], result
        if file_size > self._soft_file_size_bytes:
            issues.append(ImportIssue(
                row=0,
                severity="WARNING",
                column="file",
                message=(
                    f"File khá lớn ({_format_size(file_size)}). "
                    "Preview/import có thể chậm hơn bình thường; hãy tránh chạy đồng thời tác vụ nặng khác."
                ),
            ))
            logger.warning(
                "event=import_file_size_warning path=%s size=%s soft_limit=%s"
                % (path.name, file_size, self._soft_file_size_bytes)
            )
        return issues, None

    def _validate_headers(self, headers: list[str]) -> list[ImportIssue]:
        issues: list[ImportIssue] = []
        normalized = [h.strip().lower() for h in headers]

        # Duplicate header names
        seen: set[str] = set()
        for h in normalized:
            if h in seen:
                issues.append(ImportIssue(
                    row=0, severity="ERROR", column=h,
                    message=f"Tên cột bị trùng: '{h}'. Không thể import.",
                ))
            seen.add(h)

        # Required columns
        header_set = set(normalized)
        for col in REQUIRED_COLUMNS:
            if col not in header_set:
                issues.append(ImportIssue(
                    row=0, severity="ERROR", column=col,
                    message=f"Thiếu cột bắt buộc: '{col}'.",
                ))
        return issues

    def _parse_row(
        self, row_num: int, row: dict[str, Any]
    ) -> tuple[ParsedQuestion | None, list[ImportIssue]]:
        issues: list[ImportIssue] = []
        g = _RowGetter(row)

        # Skip if both key fields are blank
        if not g.str("question_text") and not g.str("question_type"):
            return None, []

        # --- question_text ---
        question_text = g.str("question_text")
        if not question_text:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="question_text",
                message="question_text không được để trống.",
            ))
            return None, issues

        # --- question_type ---
        qt_raw = g.str("question_type").lower()
        if qt_raw not in QUESTION_TYPE_IMPORT_MAP:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="question_type",
                message=(
                    f"question_type không hợp lệ: '{qt_raw}'. "
                    f"Chấp nhận: {list(QUESTION_TYPE_IMPORT_MAP.keys())}."
                ),
            ))
            return None, issues
        question_type = QUESTION_TYPE_IMPORT_MAP[qt_raw]

        # --- correct_answers (presence check only; detailed check later) ---
        correct_raw = g.str("correct_answers")
        if not correct_raw:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="correct_answers",
                message="correct_answers không được để trống.",
            ))
            return None, issues

        # --- options ---
        options: dict[str, str] = {}
        for label, col_name in OPTION_COLUMNS.items():
            val = g.str(col_name)
            if val:
                options[label] = val

        # --- score ---
        score_raw = g.str("score")
        score = DEFAULT_SCORE
        if score_raw:
            try:
                score = float(score_raw)
                if score <= 0:
                    raise ValueError("score must be positive")
            except ValueError:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="score",
                    message=f"score phải là số dương; nhận được: '{score_raw}'.",
                ))
                score = DEFAULT_SCORE
        else:
            issues.append(ImportIssue(
                row=row_num, severity="INFO", column="score",
                message="score để trống; sử dụng giá trị mặc định 1.0.",
            ))

        # --- difficulty ---
        diff_raw = g.str("difficulty").lower()
        valid_diffs = {d.value for d in Difficulty}
        if diff_raw and diff_raw not in valid_diffs:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="difficulty",
                message=(
                    f"difficulty không hợp lệ: '{diff_raw}'. "
                    f"Chấp nhận: {sorted(valid_diffs)}."
                ),
            ))
            diff_raw = DEFAULT_DIFFICULTY.value
        elif not diff_raw:
            diff_raw = DEFAULT_DIFFICULTY.value

        # --- status ---
        status_raw = g.str("status").lower()
        valid_statuses = {s.value for s in QuestionStatus}
        if status_raw and status_raw not in valid_statuses:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="status",
                message=f"status không hợp lệ: '{status_raw}'.",
            ))
            status_raw = DEFAULT_STATUS.value
        elif not status_raw:
            status_raw = DEFAULT_STATUS.value

        # --- boolean flags ---
        case_sensitive, cs_err = _parse_bool(g.str("case_sensitive"), default=False)
        if cs_err:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="case_sensitive", message=cs_err,
            ))

        trim_whitespace, tw_err = _parse_bool(g.str("trim_whitespace"), default=True)
        if tw_err:
            issues.append(ImportIssue(
                row=row_num, severity="ERROR", column="trim_whitespace", message=tw_err,
            ))

        # --- tags ---
        tags_raw = g.str("tags")
        tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

        # --- optional string fields ---
        question_code = g.str("question_code") or None
        category = g.str("category") or None
        hint = g.str("hint") or None
        explanation = g.str("explanation") or None

        # --- abort if already have errors (type-specific validation needs clean state) ---
        if any(i.severity == "ERROR" for i in issues):
            return None, issues

        # --- type-specific validation ---
        type_issues = self._validate_for_type(
            row_num, question_type, question_text, correct_raw, options
        )
        issues.extend(type_issues)
        if any(i.severity == "ERROR" for i in type_issues):
            return None, issues

        # --- parse correct_answers into canonical list ---
        if question_type in (QuestionType.MULTIPLE_CHOICE, QuestionType.MULTIPLE_ANSWER):
            correct_answers = [
                t.strip().upper()
                for t in correct_raw.split(MULTI_VALUE_DELIMITER)
                if t.strip()
            ]
        else:
            # BLANK / SA: preserve original text values
            correct_answers = [
                t.strip()
                for t in correct_raw.split(MULTI_VALUE_DELIMITER)
                if t.strip()
            ]

        parsed = ParsedQuestion(
            row_number=row_num,
            question_code=question_code,
            question_text=question_text,
            question_type=question_type,
            category=category,
            difficulty=diff_raw,
            score=score,
            hint=hint,
            explanation=explanation,
            options=options,
            correct_answers=correct_answers,
            status=status_raw,
            tags=tags,
            case_sensitive=case_sensitive,
            trim_whitespace=trim_whitespace,
        )
        return parsed, issues

    def _validate_for_type(
        self,
        row_num: int,
        question_type: QuestionType,
        question_text: str,
        correct_raw: str,
        options: dict[str, str],
    ) -> list[ImportIssue]:
        issues: list[ImportIssue] = []
        present_labels = list(options.keys())
        has_placeholder = count_blank_placeholders(question_text) > 0
        has_legacy_placeholder = (
            LEGACY_BLANK_PLACEHOLDER.lower() in question_text.lower()
        )

        if has_legacy_placeholder:
            issues.append(ImportIssue(
                row=row_num,
                severity="WARNING",
                column="question_text",
                message=(
                    "Đang dùng placeholder legacy '________'. "
                    "Nên đổi sang '[[blank]]' để tuân thủ chuẩn import hiện tại."
                ),
            ))

        if question_type == QuestionType.MULTIPLE_CHOICE:
            if len(options) < 2:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="option_a",
                    message="multiple_choice cần ít nhất 2 lựa chọn (option_a, option_b, ...).",
                ))
            if has_placeholder:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="question_text",
                    message=(
                        "multiple_choice không được dùng placeholder BLANK "
                        "([[blank]] hoặc legacy ________) trong question_text."
                    ),
                ))
            for err in validate_correct_answers_for_type(
                question_type, correct_raw, present_labels
            ):
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="correct_answers", message=err,
                ))

        elif question_type == QuestionType.MULTIPLE_ANSWER:
            if len(options) < 2:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="option_a",
                    message="multiple_answer cần ít nhất 2 lựa chọn (option_a, option_b, ...).",
                ))
            if has_placeholder:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="question_text",
                    message=(
                        "multiple_answer không được dùng placeholder BLANK "
                        "([[blank]] hoặc legacy ________) trong question_text."
                    ),
                ))
            for err in validate_correct_answers_for_type(
                question_type, correct_raw, present_labels
            ):
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="correct_answers", message=err,
                ))

        elif question_type == QuestionType.BLANK:
            count = count_blank_placeholders(question_text)
            if count == 0:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="question_text",
                    message=(
                        "blank question bắt buộc phải có ít nhất một placeholder "
                        "([[blank]] hoặc legacy ________) "
                        "trong question_text."
                    ),
                ))
            if options:
                issues.append(ImportIssue(
                    row=row_num, severity="WARNING", column="option_a",
                    message=(
                        "blank question không dùng option_a đến option_f; "
                        "các cột này sẽ bị bỏ qua khi import."
                    ),
                ))

        elif question_type == QuestionType.SHORT_ANSWER:
            if has_placeholder:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="question_text",
                    message=(
                        "short_answer không được dùng placeholder BLANK "
                        "([[blank]] hoặc legacy ________) trong question_text."
                    ),
                ))
            if options:
                issues.append(ImportIssue(
                    row=row_num, severity="WARNING", column="option_a",
                    message=(
                        "short_answer không dùng option_a đến option_f; "
                        "các cột này sẽ bị bỏ qua khi import."
                    ),
                ))

        elif question_type == QuestionType.TRUE_FALSE:
            if len(options) != 2:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="option_a",
                    message="true_false cần đúng 2 lựa chọn (ví dụ option_a, option_b).",
                ))
            if has_placeholder:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="question_text",
                    message="true_false không được dùng placeholder BLANK trong question_text.",
                ))
            for err in validate_correct_answers_for_type(
                question_type, correct_raw, present_labels
            ):
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="correct_answers", message=err,
                ))

        elif question_type in (QuestionType.ESSAY, QuestionType.PROBLEM):
            if has_placeholder:
                issues.append(ImportIssue(
                    row=row_num, severity="ERROR", column="question_text",
                    message="crq không được dùng placeholder BLANK trong question_text.",
                ))
            if options:
                issues.append(ImportIssue(
                    row=row_num, severity="WARNING", column="option_a",
                    message="crq không dùng option_a đến option_f; các cột này sẽ bị bỏ qua khi import.",
                ))

        return issues


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

class _RowGetter:
    """Safe accessor for a CSV / XLSX row dict; normalises key to lowercase."""

    def __init__(self, row: dict[str, Any]) -> None:
        self._row = {str(k).strip().lower(): v for k, v in row.items()}

    def str(self, key: str) -> str:
        val = self._row.get(key.lower())
        if val is None:
            return ""
        return str(val).strip()


def _parse_bool(raw: str, *, default: bool) -> tuple[bool, str | None]:
    """Normalise a boolean string from import file.

    Returns (value, error_message_or_None).
    """
    if not raw:
        return default, None
    low = raw.lower().strip()
    if low in VALID_BOOL_TRUE:
        return True, None
    if low in VALID_BOOL_FALSE:
        return False, None
    return default, (
        f"Giá trị boolean không hợp lệ: '{raw}'. "
        "Chấp nhận: true/false/1/0/yes/no."
    )


def _format_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.1f} MB"
